from __future__ import annotations

import html
import io
import ipaddress
import os
import socket
from pathlib import Path
from typing import Sequence
from urllib.parse import unquote, urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from weasyprint import HTML, default_url_fetcher

# Absolute path so the WeasyPrint url_fetcher allowlist can compare realpaths.
FONT_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "assets", "fonts", "KantumruyPro-Variable.ttf")
)
# The single local asset reports are allowed to load (the embedded Khmer font).
_FONT_URL = Path(FONT_PATH).as_uri()


# ── Security: SSRF / local-file-read protection for WeasyPrint ──────────
#
# Report rows contain user-supplied data (participant names, addresses, …).
# WeasyPrint resolves every URL in the rendered HTML/CSS through its
# ``url_fetcher`` — including ``file://`` and ``http(s)://`` — so an attacker who
# registers a name like ``<img src="http://169.254.169.254/...">`` could trigger
# SSRF or local-file disclosure. We deny-by-default: the only resource a report
# may load is the bundled font (plus inert ``data:`` URIs). Everything else —
# any ``file://`` other than the font, any remote host, any other scheme — is
# rejected. Combined with HTML-escaping every interpolated value, this makes
# generated reports incapable of fetching arbitrary resources.


class ReportAssetBlocked(Exception):
    """Raised when a report tries to load a resource that is not allowlisted."""


def _ip_is_internal(ip: ipaddress._BaseAddress) -> bool:
    return (
        ip.is_private
        or ip.is_loopback
        or ip.is_link_local      # 169.254.0.0/16 — includes cloud metadata 169.254.169.254
        or ip.is_reserved
        or ip.is_multicast
        or ip.is_unspecified
    )


_BLOCKED_HOSTNAMES = {"localhost", "metadata", "metadata.google.internal"}


def _host_is_blocked(host: str) -> bool:
    """True if a hostname/IP resolves to a loopback/private/link-local/metadata target."""
    if not host:
        return True
    h = host.strip("[]").lower()
    if h in _BLOCKED_HOSTNAMES:
        return True
    # Literal IP?
    try:
        return _ip_is_internal(ipaddress.ip_address(h))
    except ValueError:
        pass
    # Hostname → block if ANY resolved address is internal (DNS-rebinding safe).
    try:
        infos = socket.getaddrinfo(h, None)
    except OSError:
        return True  # unresolvable → block
    return any(_ip_is_internal(ipaddress.ip_address(info[4][0])) for info in infos)


def _is_allowed_font(url: str) -> bool:
    parsed = urlparse(url)
    if (parsed.scheme or "").lower() != "file":
        return False
    try:
        requested = os.path.realpath(unquote(parsed.path))
    except (ValueError, OSError):
        return False
    return requested == os.path.realpath(FONT_PATH)


def secure_url_fetcher(url: str):
    """WeasyPrint URL fetcher: allow only the bundled font + inert data: URIs."""
    scheme = (urlparse(url).scheme or "").lower()
    if scheme == "data":
        return default_url_fetcher(url)
    if scheme == "file":
        if _is_allowed_font(url):
            return default_url_fetcher(url)
        raise ReportAssetBlocked(f"Blocked local file in report asset: {url!r}")
    if scheme in ("http", "https"):
        if _host_is_blocked(urlparse(url).hostname or ""):
            raise ReportAssetBlocked(f"Blocked SSRF / internal target in report asset: {url!r}")
        # Reports never legitimately reference remote assets.
        raise ReportAssetBlocked(f"Remote assets are not permitted in reports: {url!r}")
    raise ReportAssetBlocked(f"Blocked URL scheme {scheme!r} in report asset: {url!r}")


# ── Security: Excel/CSV formula-injection guard ────────────────────────
# A cell whose text begins with = + - @ (or tab/CR) is executed as a formula
# by spreadsheet apps. User-supplied names land in XLSX cells, so neutralize
# the leading trigger with a text-forcing apostrophe.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _excel_safe(value):
    if isinstance(value, str) and value and value[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value

# ── XLSX Renderer ──────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Kantumruy Pro", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Kantumruy Pro", size=10)
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def render_xlsx(
    columns: list[tuple[str, int]],
    rows: Sequence[dict],
    col_keys: list[str] | None = None,
) -> bytes:
    """Render rows to XLSX. columns = [(header_text, width_chars), ...].

    ``col_keys`` are the dict keys to pull each column's value from; they run
    parallel to ``columns``. When omitted, the header text itself is used as the
    key (back-compat for callers whose rows are keyed by header).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    keys = col_keys if col_keys is not None else [h for h, _ in columns]

    for col_idx, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            val = row_data.get(key)
            if val is None:
                val = ""
            val = _excel_safe(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Only the explicit grand-total row (carries no="no" key set to None)
        # gets total styling; reports without a "no" column never match.
        is_total = "no" in row_data and row_data.get("no") is None
        if is_total:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = TOTAL_FILL
                ws.cell(row=row_idx, column=col_idx).font = Font(
                    name="Kantumruy Pro", bold=True, size=10
                )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── PDF Renderer ──────────────────────────────────────────────

_KHMER_NUMS = str.maketrans("0123456789", "\u17E0\u17E1\u17E2\u17E3\u17E4\u17E5\u17E6\u17E7\u17E8\u17E9")


def _kh_num(val: int | str | None) -> str:
    if val is None:
        return ""
    return str(val).translate(_KHMER_NUMS)


PDF_HTML_TPL = """\
<!DOCTYPE html>
<html lang="km">
<head>
<meta charset="utf-8">
<style>
@page {{
  size: A4 landscape;
  margin: 15mm;
}}
@font-face {{
  font-family: "Kantumruy Pro";
  src: url("{font_url}");
}}
body {{
  font-family: "Kantumruy Pro", "Noto Sans Khmer", sans-serif;
  font-size: 9pt;
  color: #222;
}}
h1 {{
  font-size: 14pt;
  margin-bottom: 4pt;
}}
h2 {{
  font-size: 10pt;
  color: #555;
  margin-top: 0;
  margin-bottom: 12pt;
}}
table {{
  width: 100%;
  border-collapse: collapse;
}}
th {{
  background-color: #4472C4;
  color: #fff;
  padding: 5pt 4pt;
  text-align: center;
  font-size: 8pt;
  font-weight: bold;
}}
td {{
  padding: 4pt;
  border: 0.5pt solid #999;
  font-size: 8pt;
  vertical-align: top;
}}
tr.total td {{
  font-weight: bold;
  background-color: #D9E2F3;
}}
tr:nth-child(even) td {{
  background-color: #f8f8f8;
}}
tr.total:nth-child(even) td {{
  background-color: #D9E2F3;
}}
.text-right {{ text-align: right; }}
.text-center {{ text-align: center; }}
</style>
</head>
<body>
<h1>{title}</h1>
<h2>{subtitle}</h2>
<table>
<thead><tr>{headers}</tr></thead>
<tbody>{rows}</tbody>
</table>
</body>
</html>"""


def _pdf_cell(value, cls=""):
    # HTML-escape every value — report rows carry user-supplied data.
    v = value if value is not None else ""
    return f'<td class="{cls}">{html.escape(str(v))}</td>'


def build_report_html(
    title: str,
    subtitle: str,
    columns: list[str],
    rows: Sequence[dict],
    col_keys: list[str],
    numeric_indices: set[int] | None = None,
) -> str:
    """Build the (fully escaped) report HTML string. Separated from rendering so
    the escaping can be unit-tested without invoking WeasyPrint."""
    if numeric_indices is None:
        numeric_indices = set()

    header_cells = "".join(f"<th>{html.escape(str(c))}</th>" for c in columns)

    body_rows = []
    for row_data in rows:
        is_total = row_data.get("no") is None
        cells = []
        for col_idx, key in enumerate(col_keys):
            val = row_data.get(key, "")
            cls = "text-right" if col_idx in numeric_indices else ""
            if col_idx in numeric_indices and isinstance(val, (int, float)):
                val = _kh_num(int(val))
            elif val is None:
                val = ""
            cells.append(_pdf_cell(val, cls))
        tr_cls = ' class="total"' if is_total else ""
        body_rows.append(f"<tr{tr_cls}>{''.join(cells)}</tr>")

    return PDF_HTML_TPL.format(
        font_url=_FONT_URL,
        title=html.escape(str(title)),
        subtitle=html.escape(str(subtitle)),
        headers=header_cells,
        rows="\n".join(body_rows),
    )


def render_pdf(
    title: str,
    subtitle: str,
    columns: list[str],
    rows: Sequence[dict],
    col_keys: list[str],
    numeric_indices: set[int] | None = None,
) -> bytes:
    """Render rows to PDF via WeasyPrint.

    columns: display header strings
    col_keys: keys to pull from each row dict
    numeric_indices: set of column indexes whose values should use Khmer numerals

    All dynamic content is HTML-escaped (see ``build_report_html``) and external
    resource loading is locked down via ``secure_url_fetcher``.
    """
    html_str = build_report_html(title, subtitle, columns, rows, col_keys, numeric_indices)
    buf = io.BytesIO()
    HTML(string=html_str, url_fetcher=secure_url_fetcher).write_pdf(buf)
    buf.seek(0)
    return buf.read()
