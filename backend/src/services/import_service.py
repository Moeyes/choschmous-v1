"""Bulk athlete import (CHOS-406).

Parses an uploaded .xlsx of athletes and runs each row through the SAME
validation + creation path as a single registration (``validate_registration`` /
``RegisterParticipant``), so a bulk import can never bypass a business rule a
one-by-one registration enforces. Produces a per-row error report the operator
can fix and re-upload.

Two modes:
  * validate (dry-run) — pydantic + business-rule validation only, no writes.
  * commit — inserts the valid rows (each committed individually, so a bad row
    never rolls back the good ones) and still reports the rows that failed.

The shared registration context (event / org / sport / category) is supplied by
the caller; each spreadsheet row supplies only the person's fields.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date, datetime

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from sqlalchemy.ext.asyncio import AsyncSession

from app.application.participants.register import RegisterParticipant
from app.application.participants.validation import validate_registration
from src.models.user import User
from src.schemas.import_athlete import ImportReport, RowError, RowResult
from src.schemas.registration import FullRegistrationRequest

logger = logging.getLogger(__name__)

# The columns the template ships with, in order. The first four names + gender +
# dateOfBirth + phone + idDocType are required by the registration schema; the
# rest are optional (document URLs are needed only to satisfy the age/document
# rule — rows missing a required document are reported, not silently accepted).
TEMPLATE_COLUMNS: list[tuple[str, str]] = [
    ("lastNameKhmer", "Last name (Khmer) *"),
    ("firstNameKhmer", "First name (Khmer) *"),
    ("lastNameLatin", "Last name (Latin) *"),
    ("firstNameLatin", "First name (Latin) *"),
    ("gender", "Gender (Male/Female) *"),
    ("dateOfBirth", "Date of birth (YYYY-MM-DD) *"),
    ("phone", "Phone *"),
    ("idDocType", "ID document type *"),
    ("nationality", "Nationality"),
    ("address", "Address"),
    ("photoUrl", "Photo URL"),
    ("nationalityDocumentUrl", "Nationality document URL"),
    ("birthCertificateUrl", "Birth certificate URL"),
    ("nationalIdUrl", "National ID URL"),
    ("passportUrl", "Passport URL"),
]

# Map a normalized header (lowercased, separators stripped) to a request field.
# Accepts the camelCase template headers plus common synonyms.
_HEADER_ALIASES: dict[str, str] = {
    "lastnamekhmer": "lastNameKhmer",
    "khfamilyname": "lastNameKhmer",
    "familynamekhmer": "lastNameKhmer",
    "firstnamekhmer": "firstNameKhmer",
    "khgivenname": "firstNameKhmer",
    "givennamekhmer": "firstNameKhmer",
    "lastnamelatin": "lastNameLatin",
    "enfamilyname": "lastNameLatin",
    "familynamelatin": "lastNameLatin",
    "firstnamelatin": "firstNameLatin",
    "engivenname": "firstNameLatin",
    "givennamelatin": "firstNameLatin",
    "gender": "gender",
    "sex": "gender",
    "dateofbirth": "dateOfBirth",
    "dob": "dateOfBirth",
    "birthdate": "dateOfBirth",
    "phone": "phone",
    "phonenumber": "phone",
    "iddoctype": "idDocType",
    "iddocumenttype": "idDocType",
    "documenttype": "idDocType",
    "nationality": "nationality",
    "address": "address",
    "photourl": "photoUrl",
    "photo": "photoUrl",
    "nationalitydocumenturl": "nationalityDocumentUrl",
    "birthcertificateurl": "birthCertificateUrl",
    "birthcertificate": "birthCertificateUrl",
    "nationalidurl": "nationalIdUrl",
    "nationalid": "nationalIdUrl",
    "passporturl": "passportUrl",
    "passport": "passportUrl",
}


@dataclass(frozen=True)
class ImportContext:
    event_id: int
    organization_id: int
    sport_id: int
    category_id: int | None
    force: bool = False


def _normalize_header(value: object) -> str:
    return (
        str(value or "")
        .strip()
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )


def build_template_workbook() -> bytes:
    """Return an .xlsx whose header row is the recognized field keys (so a
    downloaded template uploads cleanly), with the human-readable label attached
    as a cell comment for guidance. Required fields are suffixed ``*`` in the
    label."""
    from openpyxl.comments import Comment

    wb = Workbook()
    ws = wb.active
    ws.title = "athletes"
    ws.append([key for key, _ in TEMPLATE_COLUMNS])
    for col_idx, (_, label) in enumerate(TEMPLATE_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx).comment = Comment(label, "MoEYS")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class BulkAthleteImporter:
    def __init__(self, db: AsyncSession, current_user: User):
        self.db = db
        self.current_user = current_user

    def _parse(self, file_bytes: bytes) -> list[tuple[int, dict]]:
        """Parse the workbook into ``(excel_row_number, {field: value})`` pairs.
        Raises HTTPException(400) for an unreadable file or missing header."""
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(
                status_code=400, detail="Could not read the file as an .xlsx workbook."
            )
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        try:
            header = next(rows_iter)
        except StopIteration:
            raise HTTPException(status_code=400, detail="The file is empty.")

        col_map: dict[int, str] = {}
        for idx, cell in enumerate(header):
            target = _HEADER_ALIASES.get(_normalize_header(cell))
            if target:
                col_map[idx] = target
        if not col_map:
            raise HTTPException(
                status_code=400,
                detail="No recognized columns. Download the template and keep its "
                "header row.",
            )

        parsed: list[tuple[int, dict]] = []
        for offset, raw in enumerate(rows_iter, start=2):  # header is row 1
            values: dict = {}
            for idx, target in col_map.items():
                if idx < len(raw):
                    values[target] = raw[idx]
            if not _has_any_value(values):
                continue  # skip fully blank rows
            parsed.append((offset, _clean_row(values)))
        return parsed

    def _to_request(self, row: dict, ctx: ImportContext) -> FullRegistrationRequest:
        payload = {
            **row,
            "eventId": ctx.event_id,
            "organizationId": ctx.organization_id,
            "sportId": ctx.sport_id,
            "categoryId": ctx.category_id,
            "role": "athlete",
            "force": ctx.force,
        }
        # populate_by_name=True on the schema lets the camelCase aliases bind.
        return FullRegistrationRequest.model_validate(payload)

    async def run(
        self, file_bytes: bytes, ctx: ImportContext, *, commit: bool
    ) -> ImportReport:
        parsed = self._parse(file_bytes)
        results: list[RowResult] = []
        created = 0

        for excel_row, row in parsed:
            errors: list[RowError] = []
            data: FullRegistrationRequest | None = None
            try:
                data = self._to_request(row, ctx)
            except ValidationError as exc:
                errors = _errors_from_pydantic(exc)

            if data is not None:
                try:
                    if commit:
                        result = await RegisterParticipant(self.db).execute(
                            data, self.current_user
                        )
                        results.append(
                            RowResult(
                                row=excel_row,
                                ok=True,
                                enroll_id=result.get("enroll_id"),
                            )
                        )
                        created += 1
                        continue
                    else:
                        await validate_registration(self.db, data)
                        results.append(RowResult(row=excel_row, ok=True))
                        continue
                except HTTPException as exc:
                    errors = _errors_from_http(exc)
                    # validate_registration may leave the session in a usable
                    # state, but be safe before the next row's queries.
                    if not commit:
                        await _safe_rollback(self.db)

            results.append(RowResult(row=excel_row, ok=False, errors=errors))

        valid = sum(1 for r in results if r.ok)
        return ImportReport(
            committed=commit,
            total=len(results),
            valid=valid,
            invalid=len(results) - valid,
            created=created,
            rows=results,
        )


def _has_any_value(values: dict) -> bool:
    return any(v not in (None, "") for v in values.values())


def _clean_row(values: dict) -> dict:
    cleaned: dict = {}
    for key, value in values.items():
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
            if value == "":
                continue
        # openpyxl returns datetimes for date cells; the schema wants a date.
        if key == "dateOfBirth" and isinstance(value, datetime):
            value = value.date()
        elif key == "dateOfBirth" and isinstance(value, date):
            value = value
        cleaned[key] = value
    return cleaned


def _errors_from_pydantic(exc: ValidationError) -> list[RowError]:
    out: list[RowError] = []
    for err in exc.errors():
        loc = ".".join(str(p) for p in err.get("loc", ()))
        out.append(
            RowError(
                field=loc or None,
                code=err.get("type"),
                message=err.get("msg", "Invalid value"),
            )
        )
    return out


def _errors_from_http(exc: HTTPException) -> list[RowError]:
    detail = exc.detail
    if isinstance(detail, dict):
        return [
            RowError(
                code=detail.get("code"),
                message=str(detail.get("message") or detail),
            )
        ]
    return [RowError(message=str(detail))]


async def _safe_rollback(db: AsyncSession) -> None:
    try:
        await db.rollback()
    except Exception:
        logger.debug("rollback during import validate failed", exc_info=True)
