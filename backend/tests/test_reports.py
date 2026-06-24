"""Tests for the reports engine (CHOS-202 async refactor).

Rendering moved off the request path into an arq worker. These tests split into:

* **Render parity** — call ``render_report_document`` (the moved pipeline)
  directly with the test session and assert the *bytes are identical in shape*
  to the old in-route render: all 8 keys produce a valid XLSX with populated
  cells, the WeasyPrint PDF renders, and real seeded participants show up. This
  is the parity proof that the logic was moved, not changed.
* **Route contract** — ``GET /reports/{key}`` now validates + authorizes +
  enqueues and returns ``202 {job_id}`` (no live worker needed for the validation
  paths); the job-status endpoint 404s on an unknown job.
"""

import io

import pytest
from openpyxl import load_workbook

from app.application.reports.render import ReportRenderError, render_report_document
from src.models.enum.user import UserRole
from tests.conftest import make_user
from tests.factories import (
    link_org_sport,
    make_category,
    make_event,
    make_org,
    make_sport,
    make_sports_event,
)


def _xlsx_values(content: bytes) -> set[str]:
    """Flatten every non-empty cell of an XLSX into a set of strings."""
    ws = load_workbook(io.BytesIO(content)).active
    out: set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None and str(cell) != "":
                out.add(str(cell))
    return out


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REPORT_KEYS = [
    "sport-list",
    "totals",
    "counts",
    "album",
    "name-list",
    "leaders",
    "coach-athlete",
    "delegation",
]


async def _event_with_sport(db_session, sport_name="បាល់ទះ"):
    event = await make_event(db_session)
    sport = await make_sport(db_session, name_kh=sport_name)
    org = await make_org(db_session)
    await make_sports_event(db_session, event, sport)
    await link_org_sport(db_session, event, sport, org)
    return event, sport, org


# ── Render parity: the moved pipeline produces the same documents ──────────


@pytest.mark.parametrize("key", REPORT_KEYS)
async def test_report_renders_xlsx(db_session, key):
    """Every one of the 8 documents must render to a valid XLSX (not raise)."""
    event, _, _ = await _event_with_sport(db_session)
    art = await render_report_document(
        db_session,
        key=key,
        event_id=event.id,
        actor=make_user(UserRole.ADMIN),
        source=None,
        fmt="xlsx",
    )
    assert art.media_type == XLSX_MIME
    assert art.content[:2] == b"PK"  # xlsx is a zip container
    assert art.filename == f"{key}_{event.id}.xlsx"


async def test_sport_list_xlsx_cells_are_populated(db_session):
    """Regression guard: render_xlsx keys cells by col_keys, not the display
    header — otherwise every data cell renders blank."""
    event, sport, _ = await _event_with_sport(db_session, sport_name="កីឡាសាកល្បង")
    art = await render_report_document(
        db_session,
        key="sport-list",
        event_id=event.id,
        actor=make_user(UserRole.ADMIN),
        source=None,
        fmt="xlsx",
    )
    ws = load_workbook(io.BytesIO(art.content)).active
    assert ws.cell(row=1, column=2).value  # header row present
    assert ws.cell(row=2, column=1).value == 1  # "no"
    assert ws.cell(row=2, column=2).value == "កីឡាសាកល្បង"  # sport_name_kh, not blank


async def test_report_pdf_renders(db_session):
    """Exercises the WeasyPrint pipeline end-to-end (Khmer font embedded)."""
    event, _, _ = await _event_with_sport(db_session)
    art = await render_report_document(
        db_session,
        key="sport-list",
        event_id=event.id,
        actor=make_user(UserRole.ADMIN),
        source=None,
        fmt="pdf",
    )
    assert art.media_type == "application/pdf"
    assert art.content[:4] == b"%PDF"


async def test_render_unknown_key_raises(db_session):
    with pytest.raises(ReportRenderError):
        await render_report_document(
            db_session,
            key="not-a-report",
            event_id=1,
            actor=make_user(UserRole.ADMIN),
            source=None,
            fmt="xlsx",
        )


async def test_render_missing_event_raises_404(db_session):
    with pytest.raises(ReportRenderError) as exc:
        await render_report_document(
            db_session,
            key="sport-list",
            event_id=999999,
            actor=make_user(UserRole.ADMIN),
            source=None,
            fmt="xlsx",
        )
    assert exc.value.code == 404


async def test_reports_populated_with_real_participants(client, db_session, as_user):
    """End-to-end: seed an event with survey-③ counts, registered athletes, a
    coach, and an organizer (via the real registration routes), then render each
    report from the moved pipeline and assert it carries that real data."""
    from src.models.participation_per_sport import participation_per_sport
    from src.models.organizer_role import OrganizerRole

    event = await make_event(db_session)
    sport = await make_sport(db_session, name_kh="កីឡាសាក")
    org = await make_org(db_session, "ខេត្តក")
    link = await link_org_sport(db_session, event, sport, org)  # survey ②
    await make_sports_event(db_session, event, sport)
    cat = await make_category(db_session, event, sport)

    # survey ③ — planned counts (drives sport-list / totals "planned")
    db_session.add(
        participation_per_sport(
            sports_Events_id=link.id,
            org_id=org.id,
            athlete_male_count=3,
            athlete_female_count=2,
            leader_male_count=1,
            leader_female_count=0,
        )
    )
    await db_session.flush()

    as_user(make_user(UserRole.ADMIN))

    athlete = {
        "eventId": event.id,
        "organizationId": org.id,
        "sportId": sport.id,
        "categoryId": cat.id,
        "role": "athlete",
        "lastNameKhmer": "ម៉េង",
        "firstNameKhmer": "សុភា",
        "lastNameLatin": "Meng",
        "firstNameLatin": "Sophea",
        "gender": "Male",
        "dateOfBirth": "2010-03-03",
        "phone": "012111222",
        "idDocType": "IDCard",
        "birthCertificateUrl": "/u/bc.jpg",
    }
    assert (await client.post("/api/v1/registration", json=athlete)).status_code == 201
    athlete_f = {
        **athlete,
        "firstNameKhmer": "ច័ន្ទនី",
        "firstNameLatin": "Channy",
        "gender": "Female",
        "phone": "012111333",
    }
    assert (await client.post("/api/v1/registration", json=athlete_f)).status_code == 201

    coach = {
        **athlete,
        "role": "leader",
        "leaderRole": "coach",
        "categoryId": None,
        "firstNameKhmer": "រតនៈ",
        "firstNameLatin": "Ratana",
        "gender": "Male",
        "dateOfBirth": "1985-06-06",
        "phone": "012111444",
        "birthCertificateUrl": None,
        "nationalIdUrl": "/u/nid.jpg",
    }
    r_coach = await client.post("/api/v1/registration", json=coach)
    assert r_coach.status_code == 201, r_coach.text

    role = OrganizerRole(name_en="Head", name_kh="ប្រធាន", active=True)
    db_session.add(role)
    await db_session.flush()
    org_reg = await client.post(
        "/api/v1/registration/organizer",
        json={
            "eventId": event.id,
            "organizationId": org.id,
            "organizerRoleId": role.id,
            "lastNameKhmer": "ខេមរា",
            "firstNameKhmer": "វិសាល",
            "lastNameLatin": "Khemara",
            "firstNameLatin": "Visal",
            "gender": "Male",
            "dateOfBirth": "1980-02-02",
            "phone": "012111555",
            "idDocType": "IDCard",
            "nationalIdPath": "/u/nid.jpg",
        },
    )
    assert org_reg.status_code == 201, org_reg.text

    admin = make_user(UserRole.ADMIN)

    async def _render(key, fmt="xlsx"):
        art = await render_report_document(
            db_session, key=key, event_id=event.id, actor=admin, source=None, fmt=fmt
        )
        return art.content

    # sport-list: planned M=3 / F=2 (cols 4,5)
    ws = load_workbook(io.BytesIO(await _render("sport-list"))).active
    assert ws.cell(row=2, column=4).value == 3
    assert ws.cell(row=2, column=5).value == 2

    # counts: actual coaches=1 (col5), athletes=2 (col6)
    wc = load_workbook(io.BytesIO(await _render("counts"))).active
    assert wc.cell(row=2, column=5).value == 1
    assert wc.cell(row=2, column=6).value == 2

    # album / name-list carry the actual people
    assert {"ម៉េង", "សុភា", "ច័ន្ទនី"} <= _xlsx_values(await _render("album"))
    assert "រតនៈ" in _xlsx_values(await _render("name-list"))  # the coach
    assert {"ខេមរា", "ប្រធាន"} <= _xlsx_values(await _render("delegation"))

    # a fully-populated PDF renders
    pdf = await _render("album", fmt="pdf")
    assert pdf[:4] == b"%PDF"


# ── Route contract: validate + authorize + enqueue ────────────────────────


async def test_report_enqueue_returns_job_id(client, db_session, as_user):
    """The route enqueues and returns 202 + a job id without rendering inline."""
    event, _, _ = await _event_with_sport(db_session)
    as_user(make_user(UserRole.ADMIN))
    resp = await client.get(f"/api/v1/reports/sport-list?event_id={event.id}&format=xlsx")
    assert resp.status_code == 202, resp.text
    body = resp.json()
    assert body["status"] == "queued"
    assert body["job_id"]


async def test_report_unknown_key_rejected(client, db_session, as_user):
    event = await make_event(db_session)
    as_user(make_user(UserRole.ADMIN))
    resp = await client.get(
        f"/api/v1/reports/not-a-report?event_id={event.id}&format=xlsx"
    )
    assert resp.status_code == 400, resp.text


async def test_report_missing_event_404(client, db_session, as_user):
    as_user(make_user(UserRole.ADMIN))
    resp = await client.get("/api/v1/reports/sport-list?event_id=999999&format=xlsx")
    assert resp.status_code == 404, resp.text


async def test_report_bad_format_rejected(client, db_session, as_user):
    event, _, _ = await _event_with_sport(db_session)
    as_user(make_user(UserRole.ADMIN))
    resp = await client.get(f"/api/v1/reports/sport-list?event_id={event.id}&format=docx")
    assert resp.status_code == 422, resp.text  # Query pattern validation


async def test_report_job_status_unknown_returns_404(client, as_user):
    as_user(make_user(UserRole.ADMIN))
    resp = await client.get("/api/v1/reports/jobs/does-not-exist")
    assert resp.status_code == 404, resp.text
