"""Security tests for report generation (P0-1):

* HTML/PDF injection — user data is escaped before WeasyPrint renders it.
* SSRF / local-file-read — the WeasyPrint url_fetcher only allows the bundled
  font and blocks file://, ftp://, loopback/private/link-local/metadata hosts.
* Excel formula injection — leading =/+/-/@ in XLSX cells are neutralized.
"""

import io

import pytest
from openpyxl import load_workbook

from src.services.report_renderers import (
    ReportAssetBlocked,
    _FONT_URL,
    _excel_safe,
    build_report_html,
    render_pdf,
    render_xlsx,
    secure_url_fetcher,
)


# ── Injection: dynamic content is HTML-escaped ─────────────────────────

def test_build_report_html_escapes_cell_values():
    rows = [{"no": 1, "name": '<img src=x onerror=alert(1)>'}]
    html = build_report_html("Title", "Sub", ["No", "Name"], rows, ["no", "name"])
    assert "<img src=x" not in html
    assert "&lt;img src=x onerror=alert(1)&gt;" in html


def test_build_report_html_escapes_table_breakout_attempt():
    rows = [{"no": 1, "name": "</td></table><h1>spoofed</h1>"}]
    html = build_report_html("T", "S", ["No", "Name"], rows, ["no", "name"])
    assert "</table><h1>spoofed" not in html
    assert "&lt;/td&gt;&lt;/table&gt;" in html


def test_build_report_html_escapes_title_subtitle_and_headers():
    html = build_report_html("<b>t</b>", "<i>s</i>", ["<u>h</u>"], [], ["h"])
    assert "<b>t</b>" not in html and "&lt;b&gt;t&lt;/b&gt;" in html
    assert "<i>s</i>" not in html and "&lt;i&gt;s&lt;/i&gt;" in html
    assert "<u>h</u>" not in html and "&lt;u&gt;h&lt;/u&gt;" in html


def test_render_pdf_with_malicious_name_does_not_crash_and_produces_pdf():
    # End-to-end: a malicious name renders safely (escaped) and the font still
    # loads through the locked-down url_fetcher.
    rows = [{"no": 1, "name": '<img src="http://169.254.169.254/latest/meta-data/">'}]
    pdf = render_pdf("R", "E", ["No", "Name"], rows, ["no", "name"], {0})
    assert pdf[:5] == b"%PDF-"


# ── SSRF / local-file-read: url_fetcher allowlist ──────────────────────

@pytest.mark.parametrize(
    "url",
    [
        "file:///etc/passwd",
        "file:///etc/shadow",
        "ftp://attacker.example/x",
        "gopher://attacker.example/x",
        "http://169.254.169.254/latest/meta-data/",   # cloud metadata
        "http://127.0.0.1:8000/internal",             # loopback
        "http://localhost/internal",                  # loopback name
        "http://10.0.0.5/internal",                   # private
        "http://192.168.1.1/internal",                # private
        "http://172.16.0.1/internal",                 # private
        "http://[::1]/internal",                      # ipv6 loopback
        "https://metadata.google.internal/x",         # GCP metadata
    ],
)
def test_secure_url_fetcher_blocks_dangerous_urls(url):
    with pytest.raises(ReportAssetBlocked):
        secure_url_fetcher(url)


def test_secure_url_fetcher_blocks_public_remote_assets():
    # Reports never reference remote assets; even a public host is rejected so a
    # name like <img src="http://attacker.com/track"> cannot exfiltrate.
    with pytest.raises(ReportAssetBlocked):
        secure_url_fetcher("http://example.com/track.png")


def test_secure_url_fetcher_allows_bundled_font():
    result = secure_url_fetcher(_FONT_URL)
    assert isinstance(result, dict)  # default_url_fetcher returns a resource dict


# ── Excel/CSV formula injection ────────────────────────────────────────

@pytest.mark.parametrize(
    "raw,expected",
    [
        ("=1+1", "'=1+1"),
        ("+1", "'+1"),
        ("-1", "'-1"),
        ("@SUM(A1)", "'@SUM(A1)"),
        ("=cmd|'/c calc'!A1", "'=cmd|'/c calc'!A1"),
        ("Sok", "Sok"),
        ("", ""),
    ],
)
def test_excel_safe_neutralizes_formula_triggers(raw, expected):
    assert _excel_safe(raw) == expected


def test_excel_safe_passes_through_non_strings():
    assert _excel_safe(5) == 5
    assert _excel_safe(None) is None


def test_render_xlsx_neutralizes_formula_in_cell():
    cols = [("Name", 20)]
    rows = [{"name": "=HYPERLINK('http://evil','x')"}]
    data = render_xlsx(cols, rows, ["name"])
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.cell(row=2, column=1).value.startswith("'=")
