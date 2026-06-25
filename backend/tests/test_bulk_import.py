"""CHOS-406: validated bulk athlete import — parsing, dry-run report, commit."""

import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select

from src.models.enroll import Enroll
from src.models.enum.event import PhaseStatus
from src.models.enum.user import UserRole
from src.services.import_service import (
    BulkAthleteImporter,
    ImportContext,
    build_template_workbook,
)
from tests.conftest import make_user
from tests.factories import (
    link_org_sport,
    make_category,
    make_event,
    make_org,
    make_sport,
    make_sports_event,
)

# Header row uses the recognized field keys (matches the downloaded template).
_HEADERS = [
    "lastNameKhmer",
    "firstNameKhmer",
    "lastNameLatin",
    "firstNameLatin",
    "gender",
    "dateOfBirth",
    "phone",
    "idDocType",
    "birthCertificateUrl",
]


def _sheet(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(_HEADERS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _valid_row(suffix: str = "1"):
    # kh given name varies per row: the soft-duplicate check keys on
    # (kh_family_name, kh_given_name, date_of_birth), so distinct people must
    # differ there to avoid being flagged DUPLICATE_SUSPECT.
    return [
        "សុខ",
        f"ដារ៉ា{suffix}",
        "Sok",
        f"Dara{suffix}",
        "Male",
        "2010-01-01",
        "012345678",
        "BirthCertificate",
        "https://x/bc.pdf",
    ]


async def _setup(db):
    event = await make_event(db, registration=PhaseStatus.OPEN)
    sport = await make_sport(db)
    org = await make_org(db)
    await make_sports_event(db, event, sport)
    await link_org_sport(db, event, sport, org)
    category = await make_category(db, event, sport)
    return event, sport, org, category


def _ctx(event, sport, org, category, *, force=False):
    return ImportContext(
        event_id=event.id,
        organization_id=org.id,
        sport_id=sport.id,
        category_id=category.id,
        force=force,
    )


# ── Template ─────────────────────────────────────────────────────────────────


def test_template_has_field_key_headers():
    content = build_template_workbook()
    wb = load_workbook(io.BytesIO(content))
    header = [c.value for c in next(wb.active.iter_rows())]
    assert "lastNameKhmer" in header
    assert "dateOfBirth" in header


# ── Parsing errors ───────────────────────────────────────────────────────────


async def test_unreadable_file_is_400(db_session):
    importer = BulkAthleteImporter(db_session, make_user(UserRole.ADMIN))
    from fastapi import HTTPException

    with pytest.raises(HTTPException) as ei:
        importer._parse(b"not an xlsx")
    assert ei.value.status_code == 400


async def test_unrecognized_headers_is_400(db_session):
    wb = Workbook()
    wb.active.append(["foo", "bar"])
    buf = io.BytesIO()
    wb.save(buf)
    importer = BulkAthleteImporter(db_session, make_user(UserRole.ADMIN))
    from fastapi import HTTPException

    with pytest.raises(HTTPException) as ei:
        importer._parse(buf.getvalue())
    assert ei.value.status_code == 400


# ── Dry-run validate ─────────────────────────────────────────────────────────


async def test_validate_reports_valid_and_invalid_rows(db_session):
    event, sport, org, category = await _setup(db_session)
    user = make_user(UserRole.ORGANIZATION, organization_id=org.id)

    # Row 1: valid. Row 2: minor with no birth certificate → DOCUMENT_REQUIRED.
    invalid = _valid_row("2")
    invalid[-1] = None  # drop birthCertificateUrl
    data = _sheet([_valid_row("1"), invalid])

    report = await BulkAthleteImporter(db_session, user).run(
        data, _ctx(event, sport, org, category), commit=False
    )
    assert report.committed is False
    assert report.total == 2
    assert report.valid == 1
    assert report.invalid == 1
    assert report.created == 0

    bad = next(r for r in report.rows if not r.ok)
    assert bad.row == 3  # header=1, valid=2, invalid=3
    assert bad.errors
    # nothing was written on a dry-run
    count = (
        await db_session.execute(select(func.count()).select_from(Enroll))
    ).scalar()
    assert count == 0


async def test_validate_flags_pydantic_errors(db_session):
    event, sport, org, category = await _setup(db_session)
    user = make_user(UserRole.ORGANIZATION, organization_id=org.id)

    bad = _valid_row("1")
    bad[4] = None  # drop gender (required by the schema)
    data = _sheet([bad])

    report = await BulkAthleteImporter(db_session, user).run(
        data, _ctx(event, sport, org, category), commit=False
    )
    assert report.invalid == 1
    assert any("gender" in (e.field or "") for e in report.rows[0].errors)


# ── Commit ───────────────────────────────────────────────────────────────────


async def test_commit_inserts_valid_rows(db_session):
    event, sport, org, category = await _setup(db_session)
    user = make_user(UserRole.ORGANIZATION, organization_id=org.id)
    data = _sheet([_valid_row("1"), _valid_row("2")])

    report = await BulkAthleteImporter(db_session, user).run(
        data, _ctx(event, sport, org, category), commit=True
    )
    assert report.committed is True
    assert report.created == 2
    assert report.valid == 2

    count = (
        await db_session.execute(select(func.count()).select_from(Enroll))
    ).scalar()
    assert count == 2


async def test_commit_skips_blank_rows(db_session):
    event, sport, org, category = await _setup(db_session)
    user = make_user(UserRole.ORGANIZATION, organization_id=org.id)
    data = _sheet([_valid_row("1"), [None] * len(_HEADERS)])

    report = await BulkAthleteImporter(db_session, user).run(
        data, _ctx(event, sport, org, category), commit=True
    )
    # the fully-blank row is skipped, not reported as an error
    assert report.total == 1
    assert report.created == 1
